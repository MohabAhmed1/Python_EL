
import openpyxl
import random
import os
import pandas as pd
class Employee:
    
    def __init__(self):
        self.employee_lst = []
        sheet_headers = []
        self.counter = 0
        if os.path.exists("employee_data.xlsx"):
            self.workbook = openpyxl.load_workbook("employee_data.xlsx")
        else:
            self.workbook = openpyxl.Workbook()

        self.sheet = self.workbook.active
        self.sheet.title = "Employee Data"

        if self.sheet.max_row == 1 :
            sheet_headers = ['Name','ID','Job','Salary']
            self.sheet.append(sheet_headers)
    def Menu(self):
        while True:
            print('-----------------------------------------')
            print('Welcome to the Employees simple database:')
            print('Options Menu:')
            print('1.Add employee')
            print('2.Remove employee')
            print('3.Update employee')
            print('4.Print employee data')
            print('5.Exit')
            self.option = input('Choose an option:') 
            if (self.option == '1'):
                self.Add_Employee()
            elif(self.option == '2'):
                self.Remove_Employee_data()
            elif(self.option == '3'):
                self.Update_Employee()
            elif(self.option == '4'):
                self.Print_Employee_data()
            elif(self.option == '5'):
                break
            else:
                print('Wrong Entry !!')
    def Add_Employee(self):
        self.counter +=1
        self.name = input('Enter the Employee name: ')
        self.ID = str(f'{self.counter}{random.randint(1, 200)}')
        self.job = input('Enter the Employee job: ')
        self.salary = input('Enter the Employee salary: ')
        try:
            float(self.salary)
            # Create a list for the current employee data
            self.employee_lst = [self.name, self.ID, self.job, self.salary]
                    
            self.sheet.append(self.employee_lst)
            
            self.workbook.save("employee_data.xlsx")  # Save changes to workbook
            print("Employee data added successfully.")
        except ValueError:
            print("The Salary must be float !!")
        
    def Remove_Employee_data(self):
        self.name = input('Enter the Employee name to be removed:')
        df = pd.read_excel("employee_data.xlsx")
        #determine all the rows except the row to be deleted
        matching_rows = df.loc[df['Name'] != self.name]
        if(df.loc[df['Name'] == self.name].empty):
            print("Error !! Employee name not found")
        else:
            matching_rows.to_excel("employee_data.xlsx", index=False) 
            print("Employee data Removed successfully.")

    def Update_Employee(self):
        self.name = input('Enter the Employee name to be updated:')
    
        df = pd.read_excel("employee_data.xlsx")
        #getting the required row to be updated
        row_to_update = df.loc[df['Name'] == self.name]
        if (row_to_update.empty):
            print("Error !! Employee name not found")
        else:
            self.job = input('Enter the updated Employee job: ')
            self.salary = input('Enter the updated Employee salary: ')
            try:
                float(self.salary)
                #getting the rest of the rows 
                rest_of_rows = df.loc[df['Name'] != self.name]
                row_to_update['Job']=self.job
                row_to_update['Salary']=self.salary
                #concatenate the two dataframes two print the database with unupdated data and updated data
                all_rows = pd.concat([rest_of_rows,row_to_update ], ignore_index=True)
                all_rows.to_excel("employee_data.xlsx", index=False)
                print("Employee data Updated successfully.")
            except ValueError:
                print("The Salary must be float !!")

    def Print_Employee_data(self):
        self.name = input('Enter the Employee name to print its data:')
        df = pd.read_excel("employee_data.xlsx")
        row_to_print = df.loc[df['Name'] == self.name]
        if(row_to_print.empty):
            print("Error !! Employee name not found")
        else:
            print(row_to_print)
            print("Employee data printed successfully.")
        
    def __del__(self):
        print("Thanks for using our database..")



    
e = Employee()
e.Menu()

